import streamlit as st
import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use("Agg")
import io
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression, LogisticRegression
from sklearn.ensemble import RandomForestClassifier, RandomForestRegressor
from sklearn.tree import DecisionTreeClassifier
from sklearn.svm import SVC
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.metrics import (
    mean_squared_error, r2_score,
    confusion_matrix, classification_report,
    accuracy_score, precision_score, recall_score, f1_score
)
from sklearn.cluster import KMeans, DBSCAN
from sklearn.decomposition import PCA
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ─── PAGE CONFIG ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="ML Workbench",
    page_icon="🤖",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── CUSTOM CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
        padding: 2rem;
        border-radius: 12px;
        margin-bottom: 1.5rem;
        text-align: center;
    }
    .main-header h1 { color: #e94560; margin: 0; font-size: 2.5rem; }
    .main-header p  { color: #a8b2d8; margin: 0.5rem 0 0; font-size: 1.1rem; }

    .metric-card {
        background: linear-gradient(135deg, #1a1a2e, #16213e);
        border: 1px solid #0f3460;
        border-radius: 10px;
        padding: 1.2rem;
        text-align: center;
        margin-bottom: 1rem;
    }
    .metric-card .value { color: #e94560; font-size: 1.8rem; font-weight: 700; }
    .metric-card .label { color: #a8b2d8; font-size: 0.85rem; margin-top: 0.3rem; }

    .section-title {
        color: #e94560;
        font-size: 1.2rem;
        font-weight: 600;
        border-left: 4px solid #e94560;
        padding-left: 0.8rem;
        margin: 1.5rem 0 1rem;
    }
    .stSelectbox label, .stRadio label, .stSlider label { color: #a8b2d8 !important; }
    .info-box {
        background: #0f3460;
        border-radius: 8px;
        padding: 1rem;
        color: #a8b2d8;
        font-size: 0.9rem;
        margin: 0.5rem 0;
    }
</style>
""", unsafe_allow_html=True)

# ─── HEADER ───────────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
    <h1>🤖 ML Workbench</h1>
    <p>Interactive Machine Learning — Train, Evaluate & Export with one click</p>
</div>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## ⚙️ Configuration")

    # ── DATA SOURCE ──────────────────────────────────────────────────────────
    st.markdown("### 📦 Data Source")
    data_source = st.radio(
        "Choose data source",
        ["Upload CSV / Excel", "Seaborn Built-in Dataset"],
        key="data_source"
    )

    df_raw = None

    if data_source == "Upload CSV / Excel":
        uploaded = st.file_uploader("Upload file", type=["csv", "xlsx", "xls"])
        if uploaded:
            try:
                if uploaded.name.endswith(".csv"):
                    df_raw = pd.read_csv(uploaded)
                else:
                    df_raw = pd.read_excel(uploaded)
                st.success(f"✅ Loaded: {df_raw.shape[0]} rows × {df_raw.shape[1]} cols")
            except Exception as e:
                st.error(f"Error reading file: {e}")
    else:
        SEABORN_DATASETS = [
            "anscombe", "attention", "car_crashes", "diamonds", "dots",
            "dowjones", "exercise", "flights", "fmri", "geyser", "healthexp",
            "iris", "mpg", "penguins", "planets", "seaice", "taxis", "tips", "titanic"
        ]
        chosen_ds = st.selectbox("Select dataset", SEABORN_DATASETS, index=SEABORN_DATASETS.index("tips"))
        try:
            df_raw = sns.load_dataset(chosen_ds)
            st.success(f"✅ {chosen_ds}: {df_raw.shape[0]} rows × {df_raw.shape[1]} cols")
        except Exception as e:
            st.error(f"Could not load '{chosen_ds}': {e}")

    st.divider()

    # ── ML TASK ──────────────────────────────────────────────────────────────
    st.markdown("### 🧠 ML Task")
    task_type = st.selectbox(
        "Task type",
        ["Supervised — Regression", "Supervised — Classification", "Unsupervised — Clustering", "Unsupervised — PCA"],
        key="task_type"
    )

    # ── ALGORITHM ────────────────────────────────────────────────────────────
    algo_map = {
        "Supervised — Regression": ["Linear Regression", "Random Forest Regressor"],
        "Supervised — Classification": ["Logistic Regression", "Random Forest Classifier", "Decision Tree", "SVM"],
        "Unsupervised — Clustering": ["K-Means", "DBSCAN"],
        "Unsupervised — PCA": ["PCA (2 Components)", "PCA (3 Components)"],
    }
    algorithm = st.selectbox("Algorithm", algo_map[task_type], key="algorithm")

    st.divider()

    # ── TEST SIZE ─────────────────────────────────────────────────────────────
    test_size = st.slider("Test split (%)", 10, 40, 20, 5) / 100
    random_state = st.number_input("Random state", value=42, step=1)

    run_btn = st.button("🚀 Train & Evaluate", use_container_width=True, type="primary")

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN AREA — DATA PREVIEW
# ═══════════════════════════════════════════════════════════════════════════════
if df_raw is None:
    st.info("👈 Choose a data source from the sidebar to get started.")
    st.stop()

# ── TABS ─────────────────────────────────────────────────────────────────────
tab_data, tab_model, tab_export = st.tabs(["📊 Data Explorer", "🤖 Model Results", "📥 Export"])

with tab_data:
    st.markdown('<div class="section-title">Dataset Preview</div>', unsafe_allow_html=True)
    st.dataframe(df_raw.head(100), use_container_width=True)

    col1, col2, col3 = st.columns(3)
    col1.metric("Rows", f"{df_raw.shape[0]:,}")
    col2.metric("Columns", df_raw.shape[1])
    col3.metric("Missing Values", int(df_raw.isnull().sum().sum()))

    c1, c2 = st.columns(2)
    with c1:
        st.markdown('<div class="section-title">Column Types</div>', unsafe_allow_html=True)
        dtype_df = pd.DataFrame({"Type": df_raw.dtypes.astype(str), "Nulls": df_raw.isnull().sum()})
        st.dataframe(dtype_df, use_container_width=True)
    with c2:
        st.markdown('<div class="section-title">Numeric Summary</div>', unsafe_allow_html=True)
        st.dataframe(df_raw.describe().round(3), use_container_width=True)

# ═══════════════════════════════════════════════════════════════════════════════
# COLUMN SELECTOR (outside tabs so it appears before run)
# ═══════════════════════════════════════════════════════════════════════════════
numeric_cols = df_raw.select_dtypes(include=np.number).columns.tolist()
all_cols = df_raw.columns.tolist()

with tab_model:
    is_supervised = task_type.startswith("Supervised")
    is_unsupervised = task_type.startswith("Unsupervised")

    st.markdown('<div class="section-title">Feature & Target Selection</div>', unsafe_allow_html=True)

    if is_supervised:
        if task_type == "Supervised — Regression":
            target_col = st.selectbox("Target column (numeric)", numeric_cols, index=len(numeric_cols)-1)
            feature_cols = st.multiselect(
                "Feature columns",
                [c for c in numeric_cols if c != target_col],
                default=[c for c in numeric_cols if c != target_col][:5]
            )
        else:
            cat_cols = df_raw.select_dtypes(include=["object", "category"]).columns.tolist()
            target_candidates = cat_cols + numeric_cols
            target_col = st.selectbox("Target column", target_candidates, index=0)
            feature_cols = st.multiselect(
                "Feature columns (numeric)",
                [c for c in numeric_cols if c != target_col],
                default=[c for c in numeric_cols if c != target_col][:5]
            )
    else:
        target_col = None
        feature_cols = st.multiselect(
            "Feature columns (numeric)",
            numeric_cols,
            default=numeric_cols[:5]
        )
        if task_type == "Unsupervised — Clustering" and algorithm == "K-Means":
            n_clusters = st.slider("Number of clusters (K)", 2, 10, 3)
        if task_type == "Unsupervised — PCA":
            color_col = st.selectbox("Color by (optional)", ["None"] + all_cols, index=0)

    if not run_btn:
        st.info("Configure settings in the sidebar and click **🚀 Train & Evaluate**.")
        st.stop()

    # ─── VALIDATE ─────────────────────────────────────────────────────────────
    if not feature_cols:
        st.error("Please select at least one feature column.")
        st.stop()

    # ─── PREPARE DATA ─────────────────────────────────────────────────────────
    if is_supervised:
        work_cols = feature_cols + [target_col]
    else:
        work_cols = feature_cols

    df_work = df_raw[work_cols].dropna()

    if is_supervised and task_type == "Supervised — Classification":
        le = LabelEncoder()
        df_work = df_work.copy()
        df_work[target_col] = le.fit_transform(df_work[target_col].astype(str))
        class_names = le.classes_
    else:
        class_names = None

    X = df_work[feature_cols]
    scaler = StandardScaler()
    X_scaled = pd.DataFrame(scaler.fit_transform(X), columns=feature_cols)

    results = {}  # store for export

    # ═══════════════════════════════════════════════════════════════════════════
    # SUPERVISED REGRESSION
    # ═══════════════════════════════════════════════════════════════════════════
    if task_type == "Supervised — Regression":
        y = df_work[target_col]
        X_train, X_test, y_train, y_test = train_test_split(X_scaled, y, test_size=test_size, random_state=int(random_state))

        model = LinearRegression() if algorithm == "Linear Regression" else RandomForestRegressor(random_state=int(random_state))
        model.fit(X_train, y_train)
        y_pred = model.predict(X_test)

        mse  = mean_squared_error(y_test, y_pred)
        rmse = np.sqrt(mse)
        r2   = r2_score(y_test, y_pred)

        st.markdown('<div class="section-title">📈 Regression Metrics</div>', unsafe_allow_html=True)
        m1, m2, m3 = st.columns(3)
        m1.markdown(f'<div class="metric-card"><div class="value">{mse:.4f}</div><div class="label">MSE</div></div>', unsafe_allow_html=True)
        m2.markdown(f'<div class="metric-card"><div class="value">{rmse:.4f}</div><div class="label">RMSE</div></div>', unsafe_allow_html=True)
        m3.markdown(f'<div class="metric-card"><div class="value">{r2:.4f}</div><div class="label">R² Score</div></div>', unsafe_allow_html=True)

        # Coefficients / Importances
        if algorithm == "Linear Regression":
            coef_df = pd.DataFrame({"Feature": feature_cols, "Coefficient": model.coef_})
        else:
            coef_df = pd.DataFrame({"Feature": feature_cols, "Importance": model.feature_importances_}).sort_values("Importance", ascending=False)

        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown('<div class="section-title">Feature Weights</div>', unsafe_allow_html=True)
            st.dataframe(coef_df.round(5), use_container_width=True)
        with col_b:
            st.markdown('<div class="section-title">Actual vs Predicted</div>', unsafe_allow_html=True)
            fig, ax = plt.subplots(figsize=(5, 4), facecolor="#1a1a2e")
            ax.set_facecolor("#16213e")
            ax.scatter(y_test, y_pred, alpha=0.6, color="#e94560", edgecolors="white", linewidths=0.3, s=40)
            lims = [min(y_test.min(), y_pred.min()), max(y_test.max(), y_pred.max())]
            ax.plot(lims, lims, "w--", linewidth=1.2)
            ax.set_xlabel("Actual", color="#a8b2d8"); ax.set_ylabel("Predicted", color="#a8b2d8")
            ax.tick_params(colors="#a8b2d8"); ax.set_title("Actual vs Predicted", color="white")
            for spine in ax.spines.values(): spine.set_color("#0f3460")
            st.pyplot(fig); plt.close()

        eval_df = pd.DataFrame({"Metric": ["MSE", "RMSE", "R²"], "Value": [mse, rmse, r2]})
        results = dict(
            X_train=pd.DataFrame(X_train, columns=feature_cols).assign(**{target_col: y_train.values}),
            X_test=pd.DataFrame(X_test, columns=feature_cols).assign(**{target_col: y_test.values}),
            coef_df=coef_df,
            eval_df=eval_df,
            task="regression"
        )

    # ═══════════════════════════════════════════════════════════════════════════
    # SUPERVISED CLASSIFICATION
    # ═══════════════════════════════════════════════════════════════════════════
    elif task_type == "Supervised — Classification":
        y = df_work[target_col]
        X_train, X_test, y_train, y_test = train_test_split(X_scaled, y, test_size=test_size, random_state=int(random_state))

        algo_obj = {
            "Logistic Regression":       LogisticRegression(max_iter=1000, random_state=int(random_state)),
            "Random Forest Classifier":  RandomForestClassifier(random_state=int(random_state)),
            "Decision Tree":             DecisionTreeClassifier(random_state=int(random_state)),
            "SVM":                       SVC(random_state=int(random_state))
        }[algorithm]

        algo_obj.fit(X_train, y_train)
        y_pred = algo_obj.predict(X_test)

        acc  = accuracy_score(y_test, y_pred)
        prec = precision_score(y_test, y_pred, average="weighted", zero_division=0)
        rec  = recall_score(y_test, y_pred, average="weighted", zero_division=0)
        f1   = f1_score(y_test, y_pred, average="weighted", zero_division=0)
        cm   = confusion_matrix(y_test, y_pred)

        st.markdown('<div class="section-title">🎯 Classification Metrics</div>', unsafe_allow_html=True)
        mc1, mc2, mc3, mc4 = st.columns(4)
        for col, val, lbl in [(mc1, acc, "Accuracy"), (mc2, prec, "Precision"), (mc3, rec, "Recall"), (mc4, f1, "F1 Score")]:
            col.markdown(f'<div class="metric-card"><div class="value">{val:.4f}</div><div class="label">{lbl}</div></div>', unsafe_allow_html=True)

        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown('<div class="section-title">Confusion Matrix</div>', unsafe_allow_html=True)
            fig, ax = plt.subplots(figsize=(5, 4), facecolor="#1a1a2e")
            ax.set_facecolor("#16213e")
            labels = class_names if class_names is not None else [str(c) for c in sorted(y.unique())]
            im = ax.imshow(cm, cmap="RdBu_r")
            ax.set_xticks(range(len(labels))); ax.set_yticks(range(len(labels)))
            ax.set_xticklabels(labels, color="#a8b2d8", rotation=45, ha="right")
            ax.set_yticklabels(labels, color="#a8b2d8")
            for i in range(cm.shape[0]):
                for j in range(cm.shape[1]):
                    ax.text(j, i, str(cm[i, j]), ha="center", va="center", color="white", fontsize=11, fontweight="bold")
            ax.set_xlabel("Predicted", color="#a8b2d8"); ax.set_ylabel("Actual", color="#a8b2d8")
            ax.set_title("Confusion Matrix", color="white")
            plt.colorbar(im, ax=ax)
            st.pyplot(fig); plt.close()
        with col_b:
            st.markdown('<div class="section-title">Classification Report</div>', unsafe_allow_html=True)
            report = classification_report(y_test, y_pred, output_dict=True, zero_division=0)
            report_df = pd.DataFrame(report).transpose().round(4)
            st.dataframe(report_df, use_container_width=True)

        # Feature importance
        if hasattr(algo_obj, "feature_importances_"):
            coef_df = pd.DataFrame({"Feature": feature_cols, "Importance": algo_obj.feature_importances_}).sort_values("Importance", ascending=False)
        elif hasattr(algo_obj, "coef_"):
            coef_arr = algo_obj.coef_[0] if algo_obj.coef_.ndim > 1 else algo_obj.coef_
            coef_df = pd.DataFrame({"Feature": feature_cols, "Coefficient": coef_arr})
        else:
            coef_df = pd.DataFrame({"Feature": feature_cols, "Note": ["N/A"]*len(feature_cols)})

        cm_df = pd.DataFrame(cm,
            index=[f"Actual: {l}" for l in labels],
            columns=[f"Pred: {l}" for l in labels])
        eval_df = pd.DataFrame({
            "Metric": ["Accuracy", "Precision (weighted)", "Recall (weighted)", "F1 Score (weighted)"],
            "Value":  [acc, prec, rec, f1]
        })
        results = dict(
            X_train=pd.DataFrame(X_train, columns=feature_cols).assign(**{target_col: y_train.values}),
            X_test=pd.DataFrame(X_test, columns=feature_cols).assign(**{target_col: y_test.values}),
            coef_df=coef_df,
            eval_df=eval_df,
            cm_df=cm_df,
            task="classification"
        )

    # ═══════════════════════════════════════════════════════════════════════════
    # UNSUPERVISED — CLUSTERING
    # ═══════════════════════════════════════════════════════════════════════════
    elif task_type == "Unsupervised — Clustering":
        st.markdown('<div class="section-title">🔵 Cluster Results</div>', unsafe_allow_html=True)

        if algorithm == "K-Means":
            model = KMeans(n_clusters=n_clusters, random_state=int(random_state), n_init=10)
        else:
            model = DBSCAN(eps=0.5, min_samples=5)

        labels_arr = model.fit_predict(X_scaled)
        df_clustered = df_work.copy()
        df_clustered["Cluster"] = labels_arr

        n_found = len(set(labels_arr)) - (1 if -1 in labels_arr else 0)
        st.metric("Clusters found", n_found)

        # PCA for 2D plot
        pca2 = PCA(n_components=2)
        X_2d = pca2.fit_transform(X_scaled)

        fig, ax = plt.subplots(figsize=(7, 5), facecolor="#1a1a2e")
        ax.set_facecolor("#16213e")
        palette = plt.cm.tab10
        unique_labels = sorted(set(labels_arr))
        for lbl in unique_labels:
            mask = labels_arr == lbl
            name = f"Noise" if lbl == -1 else f"Cluster {lbl}"
            color = "gray" if lbl == -1 else palette(lbl / max(1, len(unique_labels)-1))
            ax.scatter(X_2d[mask, 0], X_2d[mask, 1], label=name, alpha=0.7, s=40, color=color, edgecolors="white", linewidths=0.2)
        ax.set_xlabel("PC1", color="#a8b2d8"); ax.set_ylabel("PC2", color="#a8b2d8")
        ax.tick_params(colors="#a8b2d8")
        ax.set_title(f"{algorithm} — 2D PCA View", color="white")
        ax.legend(framealpha=0.3, labelcolor="#a8b2d8", facecolor="#0f3460")
        for spine in ax.spines.values(): spine.set_color("#0f3460")
        st.pyplot(fig); plt.close()

        st.markdown('<div class="section-title">Cluster Summary</div>', unsafe_allow_html=True)
        cluster_summary = df_clustered.groupby("Cluster")[feature_cols].mean().round(3)
        st.dataframe(cluster_summary, use_container_width=True)

        eval_df = pd.DataFrame({"Metric": ["Algorithm", "Clusters Found", "Total Points"],
                                 "Value":  [algorithm, n_found, len(labels_arr)]})
        results = dict(
            X_train=df_clustered,
            X_test=pd.DataFrame(),
            coef_df=cluster_summary.reset_index(),
            eval_df=eval_df,
            task="clustering"
        )

    # ═══════════════════════════════════════════════════════════════════════════
    # UNSUPERVISED — PCA
    # ═══════════════════════════════════════════════════════════════════════════
    elif task_type == "Unsupervised — PCA":
        n_comp = 2 if algorithm == "PCA (2 Components)" else 3
        pca = PCA(n_components=n_comp)
        X_pca = pca.fit_transform(X_scaled)

        var_ratio = pca.explained_variance_ratio_
        cumvar    = np.cumsum(var_ratio)

        st.markdown('<div class="section-title">📐 PCA Results</div>', unsafe_allow_html=True)
        for i, (v, cv) in enumerate(zip(var_ratio, cumvar)):
            c1, c2 = st.columns(2)
            c1.metric(f"PC{i+1} Explained Variance", f"{v*100:.2f}%")
            c2.metric(f"PC{i+1} Cumulative", f"{cv*100:.2f}%")

        col_a, col_b = st.columns(2)
        with col_a:
            fig, ax = plt.subplots(figsize=(5, 4), facecolor="#1a1a2e")
            ax.set_facecolor("#16213e")
            ax.bar(range(1, n_comp+1), var_ratio*100, color="#e94560", alpha=0.9, edgecolor="white")
            ax.plot(range(1, n_comp+1), cumvar*100, "w-o", linewidth=1.5)
            ax.set_xlabel("Component", color="#a8b2d8"); ax.set_ylabel("Variance (%)", color="#a8b2d8")
            ax.set_title("Scree Plot", color="white"); ax.tick_params(colors="#a8b2d8")
            for spine in ax.spines.values(): spine.set_color("#0f3460")
            st.pyplot(fig); plt.close()
        with col_b:
            c_col = color_col if "color_col" in dir() and color_col != "None" else None
            fig, ax = plt.subplots(figsize=(5, 4), facecolor="#1a1a2e")
            ax.set_facecolor("#16213e")
            if c_col and c_col in df_raw.columns:
                c_vals = df_raw.loc[df_work.index, c_col].astype("category").cat.codes
                sc = ax.scatter(X_pca[:, 0], X_pca[:, 1], c=c_vals, cmap="RdBu_r", alpha=0.7, s=30, edgecolors="white", linewidths=0.2)
            else:
                ax.scatter(X_pca[:, 0], X_pca[:, 1], color="#e94560", alpha=0.6, s=30, edgecolors="white", linewidths=0.2)
            ax.set_xlabel("PC1", color="#a8b2d8"); ax.set_ylabel("PC2", color="#a8b2d8")
            ax.set_title("PCA Biplot (PC1 vs PC2)", color="white"); ax.tick_params(colors="#a8b2d8")
            for spine in ax.spines.values(): spine.set_color("#0f3460")
            st.pyplot(fig); plt.close()

        loadings = pd.DataFrame(pca.components_.T, index=feature_cols,
                                columns=[f"PC{i+1}" for i in range(n_comp)]).round(4)
        st.markdown('<div class="section-title">Loadings Matrix</div>', unsafe_allow_html=True)
        st.dataframe(loadings, use_container_width=True)

        eval_df = pd.DataFrame({
            "Component": [f"PC{i+1}" for i in range(n_comp)],
            "Explained Variance (%)": (var_ratio*100).round(4),
            "Cumulative (%)": (cumvar*100).round(4)
        })
        results = dict(
            X_train=pd.DataFrame(X_pca, columns=[f"PC{i+1}" for i in range(n_comp)]),
            X_test=pd.DataFrame(),
            coef_df=loadings.reset_index().rename(columns={"index": "Feature"}),
            eval_df=eval_df,
            task="pca"
        )

    st.session_state["results"] = results
    st.session_state["df_raw"] = df_raw

# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT TAB
# ═══════════════════════════════════════════════════════════════════════════════
with tab_export:
    st.markdown('<div class="section-title">📥 Export Results to Excel</div>', unsafe_allow_html=True)

    if "results" not in st.session_state or not st.session_state["results"]:
        st.info("Train a model first to enable export.")
    else:
        res   = st.session_state["results"]
        df_r  = st.session_state.get("df_raw", pd.DataFrame())

        def style_header(ws, row=1, fill_color="1a1a2e", font_color="e94560"):
            hdr_fill = PatternFill("solid", fgColor=fill_color)
            hdr_font = Font(bold=True, color=font_color, name="Arial")
            thin = Side(border_style="thin", color="CCCCCC")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for cell in ws[row]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        def auto_width(ws):
            for col in ws.columns:
                max_len = max((len(str(c.value)) if c.value else 0) for c in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

        def df_to_sheet(ws, df, sheet_title=None):
            if sheet_title:
                ws.append([sheet_title])
                ws["A1"].font = Font(bold=True, size=13, color="e94560", name="Arial")
                ws.append([])
            for r in [list(df.columns)] + df.values.tolist():
                ws.append([str(v) if not isinstance(v, (int, float, type(None))) else v for v in r])
            hdr_row = 3 if sheet_title else 1
            style_header(ws, row=hdr_row)
            auto_width(ws)

        buf = io.BytesIO()
        wb = openpyxl.Workbook()

        # Sheet 1 — Raw Data
        ws1 = wb.active; ws1.title = "1 - Raw Data"
        df_to_sheet(ws1, df_r.head(10000), "Raw Data")

        # Sheet 2 — Training Set
        ws2 = wb.create_sheet("2 - Training Set")
        df_to_sheet(ws2, res.get("X_train", pd.DataFrame()), "Training Set")

        # Sheet 3 — Test Set
        ws3 = wb.create_sheet("3 - Test Set")
        test_df = res.get("X_test", pd.DataFrame())
        if test_df.empty:
            ws3.append(["N/A for unsupervised methods"])
        else:
            df_to_sheet(ws3, test_df, "Test Set")

        # Sheet 4 — Coefficients / Importances
        ws4 = wb.create_sheet("4 - Model Coefficients")
        df_to_sheet(ws4, res.get("coef_df", pd.DataFrame()), "Model Coefficients / Feature Weights")

        # Sheet 5 — Evaluation
        ws5 = wb.create_sheet("5 - Model Evaluation")
        df_to_sheet(ws5, res.get("eval_df", pd.DataFrame()), "Model Evaluation Parameters")
        if "cm_df" in res:
            row_start = ws5.max_row + 2
            ws5.cell(row=row_start, column=1, value="Confusion Matrix")
            ws5.cell(row=row_start, column=1).font = Font(bold=True, color="e94560", name="Arial")
            cm_df = res["cm_df"]
            ws5.append([""] + list(cm_df.columns))
            style_header(ws5, row=row_start+1)
            for idx_name, row_vals in zip(cm_df.index, cm_df.values):
                ws5.append([idx_name] + row_vals.tolist())
            auto_width(ws5)

        wb.save(buf)
        buf.seek(0)

        st.success("✅ Workbook ready — 5 sheets packed with your results!")
        st.download_button(
            label="⬇️ Download Excel Workbook",
            data=buf.getvalue(),
            file_name="ml_workbench_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
        st.markdown("""
        <div class="info-box">
        📋 <b>Workbook structure:</b><br>
        &nbsp;&nbsp;• Sheet 1 – Raw Data (up to 10,000 rows)<br>
        &nbsp;&nbsp;• Sheet 2 – Training Set<br>
        &nbsp;&nbsp;• Sheet 3 – Test Set<br>
        &nbsp;&nbsp;• Sheet 4 – Model Coefficients / Feature Weights / Loadings<br>
        &nbsp;&nbsp;• Sheet 5 – Model Evaluation (metrics + confusion matrix if applicable)
        </div>
        """, unsafe_allow_html=True)